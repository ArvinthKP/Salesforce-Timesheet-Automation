from asyncio import tasks
from lib2to3.pgen2 import driver
from warnings import catch_warnings
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import Workbook
import openpyxl
import datetime
import time

URL = 'https://****.my.salesforce.com/'
EXCEL_FILE_PATH = "./Tasks.xlsx"

USERNAME = ""
PASSWORD = ""


class timesheet:
    def __init__(self):
        self.projectName = None
        self.taskName = None
        self.type = None
        self.hours = None
        self.remarks = None
        self.date = None


def getChromeDriver(showBrowser: bool = False, disableExtension: bool = False, noSandBox: bool = False):
    driver_path = "./chromedriver"
    chrome_options = webdriver.ChromeOptions()
    dir_path = os.getcwd()
    chrome_options.add_argument(f'user-data-dir={dir_path}/selenium')
    if(showBrowser == True):
        chrome_options.add_argument("--headless")
    if(disableExtension == True):
        chrome_options.add_argument("--disable-extensions")
    if(noSandBox == True):
        chrome_options.add_argument("--no-sandbox")
    return webdriver.Chrome(options=chrome_options, executable_path=driver_path)


def browseSite(browser, link: str, timeout: int = 60, sleeptime: int = 3):
    browser.set_page_load_timeout(timeout)
    browser.get(link)
    time.sleep(sleeptime)


def refreshBrowser(driver, sleepTIme: int = 5):
    actions = ActionChains(driver)
    actions.key_down(Keys.CONTROL)
    actions.send_keys(Keys.F5)
    actions.key_up(Keys.CONTROL)
    time.sleep(sleepTIme)


def setElementValue(driver, xpath, value, sendKeys: bool = False, index: int = 0, sleepTime: int = 1):
    elements = driver.find_elements(by=By.XPATH, value=xpath)
    inputElement = elements[index]
    if sendKeys == False:
        driver.execute_script(
            "arguments[0].value='" + value + "';", inputElement)
    else:
        action_chain = ActionChains(driver)
        action_chain.double_click(inputElement)
        action_chain.send_keys(value)
        action_chain.perform()
    time.sleep(sleepTime)


def dropDownSelect(driver, xpath, value, sleepTime: int = 1):
    time.sleep(sleepTime)
    el = driver.find_element(by=By.XPATH, value=xpath)
    for option in el.find_elements(by=By.TAG_NAME, value="option"):
        if option.text == value:
            option.click()
            break


def clickElement(driver, path, index: int = 0, sleepTime: int = 1):
    if(index == -100):
        driver.find_element_by_id(path).click()
    elif (index == -200):
        driver.find_element(by=By.XPATH, value=path).click()
    else:
        driver.find_elements(by=By.XPATH, value=path)[index].click()
    time.sleep(sleepTime)


def login(userName: str = USERNAME, password: str = PASSWORD):

    # Login is not automated properly
    setElementValue(driver, "//input[@id='username']", userName)
    setElementValue(driver, "//input[@id='password']", password, True)
    clickElement(driver, "//input[@id='Login']")
    WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable((By.ID, "save"))
    )
    time.sleep(20)
    clickElement(driver, "//input[@id='save']")


# Loading data from Excel

excelWorkBook = openpyxl.load_workbook(EXCEL_FILE_PATH)
excelSheet = excelWorkBook.active
maxcols = excelSheet.max_column
maxrows = excelSheet.max_row

taskcount = 0
allTasks = []

for r in range(2, maxrows+1):
    newTask = timesheet()
    newTask.projectName = excelSheet.cell(row=r, column=1).value
    newTask.taskName = excelSheet.cell(row=r, column=2).value
    newTask.type = excelSheet.cell(row=r, column=3).value
    newTask.hours = excelSheet.cell(row=r, column=4).value
    newTask.remarks = excelSheet.cell(row=r, column=5).value
    newTask.date = excelSheet.cell(row=r, column=6).value
    allTasks.append(newTask)


allDates = {data.date for data in allTasks}

driver = getChromeDriver()
browseSite(driver, URL)

# Time for Logging into the salesforce portal

time.sleep(100)

# login()


for selDate in allDates:

    daysTasks = [data for data in allTasks if data.date == selDate]
    print("Date ", selDate.strftime('%Y-%m-%d'))
    clickElement(
        driver, "//button[contains(@name,'"+selDate.strftime('%Y-%m-%d')+"')] ")

    for task in daysTasks:
        dropDownSelect(
            driver, "//select[contains(@name,'proId')]", task.projectName)
        dropDownSelect(
            driver, "//select[contains(@name,'taskId')]", task.taskName, 3)
        dropDownSelect(driver, "//select[contains(@name,'typeId')]", task.type)
        setElementValue(driver, "//input[@name='inputdec']", str(task.hours))
        setElementValue(driver, "//input[@name='inputrem']", task.remarks)
        clickElement(
            driver, "//button[@type='button' and contains(., 'Save')]", -200)
        time.sleep(5)
